import models

import os
from datetime import datetime
from database import database
from sqlalchemy import and_
from models import users, investments_items, investments_history, investments_in_out, categories, key_rate
import schemas

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from exeptions import CategoryInUse, CategoryNotFound, InvestmentNotFound, KeyRateNotFound


async def get_user(user_id: int | None = None,
                   username: str | None = None,
                   email: str | None = None) -> schemas.UserInDB:
    """Get user by user_id, username or email from DB"""
    result: schemas.UserInDB | None = None
    if user_id:
        result = await database.fetch_one(users.select().where(users.c.id == user_id))
    elif username:
        result = await database.fetch_one(users.select().where(users.c.username == username))
    elif email:
        result = await database.fetch_one(users.select().where(users.c.email == email))
    return result


async def create_user(user: schemas.UserCreate, hashed_password: str) -> schemas.User:
    """Create new user with hashed password in DB"""
    db_user = users.insert().values(username=user.username,
                                    email=user.email,
                                    hashed_password=hashed_password,
                                    is_active=user.is_active)
    user_id = await database.execute(db_user)
    return schemas.User(**user.dict(), id=user_id)


async def create_user_investment_item(investment: schemas.InvestmentCreate, user_id: int) -> schemas.InvestmentInDB:
    """Create new investment in DB"""
    query = investments_items.insert().values(**investment.dict(), owner_id=user_id)
    investment_id = await database.execute(query)
    return schemas.InvestmentInDB(**investment.dict(), id=investment_id, owner_id=user_id)


async def get_user_investment_items(user_id: int) -> schemas.InvestmentUser:
    """Get user investments by user_id from DB"""
    list_investments = await database.fetch_all(investments_items.select()
                                                .where(investments_items.c.owner_id == user_id))
    return schemas.InvestmentUser(**{"investments": [dict(result) for result in list_investments]})


async def update_user_investment_item(investment: schemas.InvestmentOut, user_id: int) -> schemas.Result:
    """Update user investment in DB (description, category_id)"""
    investment_in_db = await database.fetch_one(investments_items.select()
                                                .where(and_(investments_items.c.id == investment.id,
                                                            investments_items.c.owner_id == user_id)))
    if investment_in_db:
        query = investments_items.update().where(and_(investments_items.c.id == investment.id,
                                                      investments_items.c.owner_id == user_id))\
            .values(description=investment.description, category_id=investment.category_id)
        await database.execute(query)
        return schemas.Result(**{"result": "investment updated"})
    else:
        raise InvestmentNotFound


async def delete_user_investment_item(investment_id: int, user_id: int) -> schemas.Result:
    """Delete user investment by id from DB"""
    query = investments_items.select().where(and_(investments_items.c.id == investment_id,
                                                  investments_items.c.owner_id == user_id))
    result = await database.fetch_one(query)
    if result:
        query = investments_items.delete().where(and_(investments_items.c.id == investment_id,
                                                      investments_items.c.owner_id == user_id))
        await database.execute(query)
        return schemas.Result(**{"result": "investments deleted"})
    else:
        raise InvestmentNotFound


async def create_user_category(category: schemas.CategoryCreate, user_id: int) -> schemas.CategoryInDB:
    """Create new category for user in DB"""
    query = categories.insert().values(**category.dict(), owner_id=user_id)
    category_id = await database.execute(query)
    return schemas.CategoryInDB(**category.dict(), id=category_id, owner_id=user_id)


async def get_user_categories(user_id: int) -> schemas.CategoryUser:
    """Get categories for user from DB"""
    list_categories = await database.fetch_all(categories.select().where(categories.c.owner_id == user_id))
    return schemas.CategoryUser(**{"categories": [dict(result) for result in list_categories]})


async def update_user_category(category: schemas.CategoryOut, user_id: int) -> schemas.Result:
    """Update category from DB"""
    query = categories.select().where(and_(categories.c.id == category.id,
                                           categories.c.owner_id == user_id))
    categories_found = await database.fetch_all(query)
    if categories_found:
        query = categories.update().where((categories.c.id == category.id)).values(category=category.category)
        await database.execute(query)
        return schemas.Result(**{"result": "category updated"})
    else:
        raise CategoryNotFound


async def delete_user_category(category_id: int, user_id: int) -> schemas.Result:
    """Delete unused category from DB"""
    query = investments_items.select().where(and_(investments_items.c.category_id == category_id,
                                                  investments_items.c.owner_id == user_id))
    categories_found = await database.fetch_all(query)
    if categories_found:
        raise CategoryInUse
    else:
        query = categories.select().where(and_(categories.c.id == category_id, categories.c.owner_id == user_id))
        result = await database.execute(query)
        if result:
            query = categories.delete().where(and_(categories.c.id == category_id, categories.c.owner_id == user_id))
            await database.execute(query)
            result = schemas.Result(**{"result": "category deleted"})
        else:
            raise CategoryNotFound
    return result


async def user_investment_exist(investment_id: int, user_id: int) -> bool:
    """Check exist user investment history in DB"""
    return await database.fetch_one(investments_items.select().where(and_(investments_items.c.id == investment_id,
                                                                          investments_items.c.owner_id == user_id)))


async def create_user_investment_history(investment: schemas.HistoryCreate, user_id: int) -> schemas.HistoryInDB:
    """Create new investment history in DB"""
    if await user_investment_exist(investment_id=investment.investment_id, user_id=user_id):
        query = investments_history.insert().values(**investment.dict())
        investment_id = await database.execute(query)
        return schemas.HistoryInDB(**investment.dict(), id=investment_id)
    else:
        raise InvestmentNotFound


async def get_user_investment_history(user_id: int, investment_id: int) -> schemas.HistoryUser:
    """Get user investments history by user_id from DB"""
    if await user_investment_exist(investment_id=investment_id, user_id=user_id):
        list_investment_history = await database.fetch_all(investments_history.select()
                                                           .where(investments_history.c.investment_id == investment_id))
        return schemas.HistoryUser(**{"history": [dict(result) for result in list_investment_history]})
    else:
        raise InvestmentNotFound


async def update_user_investment_history(investment: schemas.HistoryOut, user_id: int) -> schemas.Result:
    """Update user investment history in DB (date, sum)"""
    if await user_investment_exist(investment_id=investment.investment_id, user_id=user_id):
        query = investments_history.update().where((investments_history.c.id == investment.id))\
            .values(date=investment.date, sum=investment.sum)
        await database.execute(query)
        return schemas.Result(**{"result": "investment history updated"})
    else:
        raise InvestmentNotFound


async def delete_user_investment_history(investment_history_id: int, user_id: int) -> schemas.Result:
    """Delete user investment history by id from DB"""
    investment_history_in_db = await database.fetch_one(investments_history
                                             .select().where(investments_history.c.id == investment_history_id))
    if investment_history_in_db and await user_investment_exist(investment_id=investment_history_in_db.investment_id,
                                                                user_id=user_id):
        query = investments_history.delete()\
            .where(and_(investments_history.c.id == investment_history_id,
                        investments_history.c.investment_id == investment_history_in_db.investment_id))
        await database.execute(query)
        return schemas.Result(**{"result": "investments history item deleted"})
    else:
        raise InvestmentNotFound


async def create_user_investment_inout(investment: schemas.InOutCreate, user_id: int) -> schemas.InOutInDB:
    """Create new investment in/out in DB"""
    if await user_investment_exist(investment_id=investment.investment_id, user_id=user_id):
        query = investments_in_out.insert().values(**investment.dict())
        investment_id = await database.execute(query)
        return schemas.InOutInDB(**investment.dict(), id=investment_id)
    else:
        raise InvestmentNotFound


async def get_user_investment_inout(user_id: int, investment_id: int) -> schemas.InOutUser:
    """Get user investments in/out by user_id from DB"""
    if await user_investment_exist(investment_id=investment_id, user_id=user_id):
        list_investment_in_out = await database.fetch_all(investments_in_out.select()
                                                           .where(investments_in_out.c.investment_id == investment_id))
        return schemas.InOutUser(**{"in_out": [dict(result) for result in list_investment_in_out]})
    else:
        raise InvestmentNotFound


async def update_user_investment_inout(investment: schemas.InOutOut, user_id: int) -> schemas.Result:
    """Update user investment in/out in DB (date, sum)"""
    if await user_investment_exist(investment_id=investment.investment_id, user_id=user_id):
        query = investments_in_out.update().where((investments_in_out.c.id == investment.id))\
            .values(date=investment.date, description=investment.description, sum=investment.sum)
        await database.execute(query)
        return schemas.Result(**{"result": "investment in/out updated"})
    else:
        raise InvestmentNotFound


async def delete_user_investment_inout(investment_in_out_id: int, user_id: int) -> schemas.Result:
    """Delete user investment in/out by id from DB"""
    investment_in_out_in_db = await database.fetch_one(investments_in_out
                                             .select().where(investments_in_out.c.id == investment_in_out_id))
    if investment_in_out_in_db and await user_investment_exist(investment_id=investment_in_out_in_db.investment_id,
                                                               user_id=user_id):
        query = investments_in_out.delete()\
            .where(and_(investments_in_out.c.id == investment_in_out_id,
                        investments_in_out.c.investment_id == investment_in_out_in_db.investment_id))
        await database.execute(query)
        return schemas.Result(**{"result": "investments in/out item deleted"})
    else:
        raise InvestmentNotFound


async def get_key_rate() -> schemas.KeyRateUser:
    """Get key rate from DB"""
    list_key_rates = await database.fetch_all(key_rate.select())
    if list_key_rates:
        return schemas.KeyRateUser(**{"key_rates": [dict(result) for result in list_key_rates]})
    else:
        raise KeyRateNotFound


async def get_investment_report_json(user_id: int) -> schemas.InvestmentReport:
    """Create investment report in json"""
    user_report = schemas.InvestmentReport()

    # get user investments
    list_investments = await database.fetch_all(investments_items.select()
                                                .where(investments_items.c.owner_id == user_id))
    list_categories = await database.fetch_all(categories.select().where(categories.c.owner_id == user_id))

    list_key_rates = await database.fetch_all(key_rate.select())

    user_categories = {}
    for category in list_categories:
        user_categories.update({category['id']: category['category']})

    for investment in list_investments:
        asset = schemas.InvestmentReportAsset()
        asset.description = investment['description']
        asset.id = investment['id']
        asset.category_id = investment['category_id']
        if user_categories:
            asset.category = user_categories[asset.category_id]

        list_investment_in_out = await database.fetch_all(investments_in_out.select()
                                                          .where(investments_in_out.c.investment_id == asset.id))

        for inout in list_investment_in_out:
            year_mon = str(inout['date'].timetuple().tm_year) + '-'\
                                  + str(inout["date"].timetuple().tm_mon).zfill(2)
            if inout['sum'] > 0:
                if year_mon in asset.sum_in:
                    asset.sum_in[year_mon] += inout['sum']
                else:
                    asset.sum_in.update({year_mon: inout['sum']})
            elif inout['sum'] < 0:
                if year_mon in asset.sum_out:
                    asset.sum_out[year_mon] += inout['sum']
                else:
                    asset.sum_out.update({year_mon: inout['sum']})

        list_investment_history = await database.fetch_all(investments_history.select()
                                                           .where(investments_history.c.investment_id == asset.id))

        for history in list_investment_history:
            year_mon = str(history['date'].timetuple().tm_year) + '-'\
                                  + str(history["date"].timetuple().tm_mon).zfill(2)
            asset.sum_fact.update({year_mon: history['sum']})

        for key_rate_item in list_key_rates:
            year_mon = str(key_rate_item['date'].timetuple().tm_year) + '-' \
                       + str(key_rate_item["date"].timetuple().tm_mon).zfill(2)
            asset.key_rates.update({year_mon: key_rate_item['key_rate']})

        dates = set(asset.sum_fact.keys()) | set(asset.sum_in.keys()) | set(asset.sum_out.keys())
        dates_sort_list = sorted(list(dates))

        tmp_dates_sort_list = []
        if len(dates_sort_list) > 0:
            year_begin = int(dates_sort_list[0][0:4])
            year_end = int(dates_sort_list[len(dates_sort_list) - 1][0:4])
            for year in range(year_begin, year_end + 1):
                for month in range(1, 13):
                    date = str(year) + "-" + str(month).zfill(2)
                    if dates_sort_list[0] <= date <= dates_sort_list[len(dates_sort_list) - 1]:
                        tmp_dates_sort_list.append(date)
            dates_sort_list = tmp_dates_sort_list

        total_sum = 0
        total_items = 0
        average_sum = 0
        average_items = 0
        last_sum_fact = 0
        deposit_index_sum = 0
        last_key_rate = 4

        for date in dates_sort_list:
            total_items += 1

            if date in asset.sum_in:
                total_sum += asset.sum_in[date]
                deposit_index_sum += asset.sum_in[date]

            if date in asset.sum_out:
                total_sum += asset.sum_in[date]
                deposit_index_sum += asset.sum_in[date]

            if date in asset.key_rates:
                deposit_index_sum += deposit_index_sum * asset.key_rates[date] / 100 / 12
                last_key_rate = asset.key_rates[date]
            else:
                deposit_index_sum += deposit_index_sum * last_key_rate / 100 / 12

            asset.sum_deposit_index[date] = int(deposit_index_sum)
            if deposit_index_sum != 0:
                asset.ratio_deposit_index[date] = int(total_sum/deposit_index_sum * 100 - 100)
            else:
                asset.ratio_deposit_index[date] = 0

            asset.sum_plan[date] = total_sum

            if date not in asset.sum_fact:
                asset.sum_fact[date] = last_sum_fact
                if deposit_index_sum != 0:
                    asset.ratio_deposit_index[date] = int(last_sum_fact / deposit_index_sum * 100 - 100)
            else:
                last_sum_fact = asset.sum_fact[date]
                if deposit_index_sum != 0:
                    asset.ratio_deposit_index[date] = int(asset.sum_fact[date] / deposit_index_sum * 100 - 100)

            asset.sum_delta_rub[date] = asset.sum_fact[date] - total_sum

            if total_sum != 0:
                asset.sum_delta_proc[date] = round((asset.sum_fact[date] - total_sum) / total_sum * 100, 1)

                average_sum += asset.sum_delta_proc[date]
                average_items += 1

                asset.sum_delta_proc_avg[date] = round(average_sum / average_items, 1)

            if total_items != 0:
                asset.sum_cashflow[date] = int((asset.sum_fact[date] - total_sum) / total_items)

        user_report.investment_report.append(asset)
    return user_report


async def get_investment_report_xlsx(user_id: int) -> str:
    """Create investment report in xlsx"""
    json_report = await get_investment_report_json(user_id)

    wb = Workbook()
    bold = Font(bold=True)

    for asset in json_report.investment_report:
        wb.create_sheet(asset.description)
        sht = wb[asset.description]
        row = column = 1

        column_dimensions = {"A": 8, "B": 13, "C": 8, "D": 12, "E": 12, "F": 14,
                             "G": 13, "H": 15, "I": 9, "J": 14, "K": 16}
    
        for col in column_dimensions.keys():
            sht.column_dimensions[col].width = column_dimensions[col]

        titles = ['Дата',
                  'Пополнение',
                  'Снятие',
                  'Сумма план',
                  'Сумма факт',
                  'Прирост руб',
                  'Прирост %',
                  'Прирост средн',
                  'Cashflow',
                  'ЕслиНаВклад',
                  'ОтклОтВклада%']

        for title in titles:
            cell = sht.cell(row=row, column=column)
            cell.value = title
            cell.font = bold
            column += 1

        row = 2
        for date in asset.sum_plan:
            column = 1
            cell = sht.cell(row=row, column=column)
            cell.value = date

            if date in asset.sum_in:
                column = 2
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_in[date]

            if date in asset.sum_out:
                column = 3
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_out[date]

            if date in asset.sum_plan:
                column = 4
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_plan[date]

            if date in asset.sum_fact:
                column = 5
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_fact[date]

            if date in asset.sum_delta_rub:
                column = 6
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_delta_rub[date]

            if date in asset.sum_delta_proc:
                column = 7
                cell = sht.cell(row=row, column=column)
                cell.value = asset.sum_delta_proc[date]

            if date in asset.sum_delta_proc_avg:
                column = 8
                cell = sht.cell(row=row, column=column)
                cell.value = cell.value = asset.sum_delta_proc_avg[date]

            if date in asset.sum_cashflow:
                column = 9
                cell = sht.cell(row=row, column=column)
                cell.value = cell.value = asset.sum_cashflow[date]

            if date in asset.sum_deposit_index:
                column = 10
                cell = sht.cell(row=row, column=column)
                cell.value = cell.value = asset.sum_deposit_index[date]

            if date in asset.ratio_deposit_index:
                column = 11
                cell = sht.cell(row=row, column=column)
                cell.value = cell.value = asset.ratio_deposit_index[date]

            row += 1

    xlsx_file = '.' + os.sep + 'static' + os.sep + 'export' + os.sep + f'investresults{user_id}.xlsx'
    sht = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(sht)
    wb.save(filename=xlsx_file)
    return xlsx_file
